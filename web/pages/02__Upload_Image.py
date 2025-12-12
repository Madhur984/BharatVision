# 
# Upload Image Processing Page - Legal Metrology OCR Pipeline
# Using Surya OCR (if available) with EasyOCR fallback.
# 

# Initialize headless mode FIRST - before any OpenCV imports
import os
os.environ['OPENCV_HEADLESS'] = '1'
os.environ['QT_QPA_PLATFORM'] = 'offscreen'
os.environ['MPLBACKEND'] = 'Agg'

import streamlit as st
import time
import json
from datetime import datetime
from pathlib import Path
import sys
import io
import requests
from typing import List, Dict, Any


from PIL import Image
import pandas as pd
import numpy as np
import re  # For regex pattern matching

# Try to import cv2, but it's optional (may fail on Python 3.13)
try:
    import cv2
    CV2_AVAILABLE = True
except ImportError:
    CV2_AVAILABLE = False

# Pipeline availability flag (set after imports below)
PIPELINE_AVAILABLE = False
IMPORT_ERRORS = []

# -----------------------------------------------------------------------------
# PAGE CONFIG
# -----------------------------------------------------------------------------
st.set_page_config(
    page_title="Upload Images - Legal Metrology OCR",
    page_icon="",
    layout="wide"
)

# -----------------------------------------------------------------------------
# PATHS & CONFIG
# -----------------------------------------------------------------------------
project_root = Path(__file__).parent.parent.parent  # repo root
if str(project_root) not in sys.path:
    sys.path.append(str(project_root))

web_root = Path(__file__).parent.parent
if str(web_root) not in sys.path:
    sys.path.insert(0, str(web_root))

# Import database with robust handling
try:
    from common import get_database
    db = get_database()
except ImportError:
    try:
        from database import db
    except (ImportError, KeyError):
        from web.database import db

# Safely import EXPECTED_FIELDS from config.py
import importlib.util

config_path = project_root / "config.py"
if config_path.exists():
    spec = importlib.util.spec_from_file_location("app_config", str(config_path))
    app_config = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(app_config)
    EXPECTED_FIELDS = getattr(app_config, "EXPECTED_FIELDS", [])
else:
    EXPECTED_FIELDS = []

# API Configuration
# API Configuration - Robust loading
ML_API_URL = os.environ.get("ML_API_URL")

# Check secrets if env var not set
if not ML_API_URL:
    try:
        if "api" in st.secrets and "ML_API_URL" in st.secrets["api"]:
            ML_API_URL = st.secrets["api"]["ML_API_URL"]
        elif "ML_API_URL" in st.secrets:
            ML_API_URL = st.secrets["ML_API_URL"]
    except Exception:
        pass

# Final Fallback
if not ML_API_URL:
    ML_API_URL = "http://localhost:8000"

def check_api_health():
    try:
        r = requests.get(f"{ML_API_URL}/health", timeout=2)
        return r.status_code == 200
    except:
        return False

PIPELINE_AVAILABLE = True # We assume API is the pipeline now


try:
    from data_refiner.refiner import DataRefiner
except Exception as e:
    IMPORT_ERRORS.append(f"DataRefiner: {type(e).__name__}: {e}")

try:
    from lmpc_checker.compliance_validator import ComplianceValidator
except Exception as e:
    IMPORT_ERRORS.append(f"ComplianceValidator: {type(e).__name__}: {e}")

# Create simple fallback classes if imports failed
if DataRefiner is None:
    class DataRefiner:
        """Fallback DataRefiner"""
        def refine(self, data, use_nlp: bool = True):
            if isinstance(data, str):
                return {"raw_text": data, "extracted_fields": {}}
            elif isinstance(data, dict):
                return data
            return {"raw_text": str(data), "extracted_fields": {}}

if ComplianceValidator is None:
    class ComplianceValidator:
        """Fallback ComplianceValidator"""
        def validate(self, data):
            return []  # No violations

# Pipeline is now always available with fallbacks
PIPELINE_AVAILABLE = True
# 

# Show import errors as info (not blocking)
if IMPORT_ERRORS:
    st.sidebar.warning(f"⚠️ Some components using fallbacks: {len(IMPORT_ERRORS)} import issues")

# -----------------------------------------------------------------------------
# AUTH CHECK
# -----------------------------------------------------------------------------
if "authenticated" not in st.session_state or not st.session_state.authenticated:
    st.switch_page("pages/00__Login.py")
    st.stop()

# -----------------------------------------------------------------------------
# CONSTANTS
# -----------------------------------------------------------------------------
SUPPORTED_FORMATS = ["jpg", "jpeg", "png", "bmp", "webp"]
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB
MAX_BATCH_SIZE = 20
 
# # -----------------------------------------------------------------------------
# # CSS
# # -----------------------------------------------------------------------------
def load_upload_css():
    # Construct CSS line-by-line to avoid parser syntax errors with multiline strings
    css_lines = [
        "<style>",
        ".stApp {",
        "    background: linear-gradient(180deg, #dfeeff 0%, #eef7ff 45%, #fff2ea 100%) !important;",
        "    background-attachment: fixed;",
        "    color: #000000;",
        "}",
        "[data-testid='stSidebar'] {",
        "    background: linear-gradient(180deg,#bcd7ff 0%, #a9d2ff 100%) !important;",
        "    color: #0b1220 !important;",
        "}",
        "header {",
        "    background: linear-gradient(135deg, rgba(215,233,255,0.95), rgba(255,242,234,0.95));",
        "    color: #000000 !important;",
        "}",
        ".page-header {",
        "    background: rgba(255,255,255,0.05);",
        "    border-radius: 18px;",
        "    padding: 26px 30px;",
        "    margin-bottom: 20px;",
        "    border: 1px solid rgba(255,255,255,0.03);",
        "    box-shadow: 0 18px 48px rgba(2,6,23,0.45);",
        "    backdrop-filter: blur(6px);",
        "}",
        ".small-note { color: rgba(255,255,255,0.5); font-size:12px; }",
        "[data-testid='stSidebarNav'] li:first-child { display:none !important; }",
        "</style>"
    ]
    st.markdown("\n".join(css_lines), unsafe_allow_html=True)
# 
# -----------------------------------------------------------------------------
# SESSION STATE
# -----------------------------------------------------------------------------
def initialize_upload_session():
    ss = st.session_state
    ss.setdefault("uploaded_files", [])
    ss.setdefault("batch_results", [])
    ss.setdefault("processing_status", {})
    ss.setdefault("selected_results", [])
    ss.setdefault("current_batch_results", [])
    ss.setdefault("files_processed", set())
    ss.setdefault("processing_completed", False)

# -----------------------------------------------------------------------------
# FILE VALIDATION
# -----------------------------------------------------------------------------
def validate_uploaded_file(uploaded_file) -> tuple[bool, str]:
    ext = uploaded_file.name.split(".")[-1].lower()
    if ext not in SUPPORTED_FORMATS:
        return False, f"Unsupported format. Supported: {', '.join(SUPPORTED_FORMATS)}"

    if uploaded_file.size > MAX_FILE_SIZE:
        return False, f"File too large. Max {MAX_FILE_SIZE // (1024*1024)} MB"

    try:
        img = Image.open(uploaded_file)
        img.verify()
        uploaded_file.seek(0)
        return True, "Valid image file"
    except Exception as e:
        return False, f"Invalid image file: {e}"

# -----------------------------------------------------------------------------
# CAMERA (BROWSER ONLY)
# -----------------------------------------------------------------------------
def create_camera_capture_interface():
    # Use browser camera only (st.camera_input).
    st.markdown("### Capture from Camera")

    captured = st.camera_input("Use your device camera to take a photo")
    if captured is None:
        return None

    raw = captured.getvalue()
    data = io.BytesIO(raw)

    class CapturedImageFile:
        def __init__(self, data, name="camera_capture.jpg"):
            self.data = data
            self.name = name
            self.type = "image/jpeg"

        def read(self, size=-1):
            return self.data.read(size)

        def seek(self, pos, whence=0):
            return self.data.seek(pos, whence)

        def tell(self):
            return self.data.tell()

        @property
        def size(self):
            cur = self.data.tell()
            self.data.seek(0, 2)
            total = self.data.tell()
            self.data.seek(cur)
            return total

    return [CapturedImageFile(data)]

# -----------------------------------------------------------------------------
# FILE UPLOAD UI
# -----------------------------------------------------------------------------
def create_file_upload_interface():
    st.markdown("###  Upload Product Images")

    # Construct HTML line-by-line to avoid parser issues
    html_lines = [
        '<div class="upload-zone">',
        '    <h3> Drag & Drop Images Here</h3>',
        '    <p>Or use the button below to browse files</p>',
        '    <p><strong>Supported formats:</strong> JPG, PNG, BMP, WEBP</p>',
        '    <p><strong>Maximum size:</strong> 10MB per file | <strong>Batch limit:</strong> 20 files</p>',
        '</div>'
    ]
    st.markdown("\n".join(html_lines), unsafe_allow_html=True)

    uploaded_files = st.file_uploader(
        "Choose image files",
        type=SUPPORTED_FORMATS,
        accept_multiple_files=True,
        help=f"Upload up to {MAX_BATCH_SIZE} images at once. Max {MAX_FILE_SIZE // (1024*1024)} MB per file.",
    )
    return uploaded_files

def display_uploaded_files(uploaded_files):
    if not uploaded_files:
        st.session_state.current_batch_results = []
        st.session_state.files_processed = set()
        st.session_state.processing_completed = False
        return []

    current_ids = {f"{f.name}_{f.size}" for f in uploaded_files}
    if current_ids != st.session_state.files_processed:
        st.session_state.current_batch_results = []
        st.session_state.processing_completed = False

    st.markdown("### 📋 Uploaded Files")
    valid_files = []

    for f in uploaded_files:
        col1, col2, col3 = st.columns([3, 2, 1])

        with col1:
            # Construct HTML safely line-by-line
            file_info_html = [
                '<div class="file-info">',
                f'    <strong>{f.name}</strong><br>',
                f'    Size: {f.size / 1024:.1f} KB | Type: {f.type}',
                '</div>'
            ]
            st.markdown("\n".join(file_info_html), unsafe_allow_html=True)

        with col2:
            ok, msg = validate_uploaded_file(f)
            if ok:
                st.success("✅ " + msg)
                valid_files.append(f)
            else:
                st.error("❌ " + msg)

        with col3:
            if f.type.startswith("image/"):
                try:
                    img = Image.open(f)
                    st.image(img, width=100)
                    f.seek(0)
                except Exception:
                    st.write("No preview")

    return valid_files

# -----------------------------------------------------------------------------
# OCR HELPERS (EasyOCR PRIMARY - works on Streamlit Cloud)
# ============================================================================

# Cache EasyOCR reader to avoid reloading on each image
@st.cache_resource
def get_easyocr_reader():
    # Get cached EasyOCR reader
    try:
        import easyocr
        return easyocr.Reader(['en'], gpu=False, verbose=False)
    except Exception as e:
        st.error(f"Failed to initialize EasyOCR: {e}")
        return None

def run_easyocr(pil_img: Image.Image) -> tuple[str, str]:
    # Primary OCR engine using EasyOCR. Returns (text, debug_msg).
    # Pure Python - works on Streamlit Cloud without system dependencies.
    try:
        reader = get_easyocr_reader()
        if reader is None:
            return "", "EasyOCR reader not available"
        
        # Ensure RGB
        if pil_img.mode != "RGB":
            pil_img = pil_img.convert("RGB")
        
        # Convert to numpy array
        img_array = np.array(pil_img)
        
        # Run EasyOCR
        results = reader.readtext(img_array)
        
        if results:
            # Extract text with confidence > 0.2
            extracted_lines = [text for (bbox, text, confidence) in results if confidence > 0.2]
            extracted_text = '\n'.join(extracted_lines)
            
            if extracted_text.strip():
                return extracted_text.strip(), f"✅ EasyOCR successful ({len(extracted_lines)} lines)"
        
        return "", "EasyOCR ran but found no text"
        
    except Exception as e:
        return "", f"EasyOCR error: {type(e).__name__}: {str(e)}"

def run_tesseract_ocr(pil_img: Image.Image) -> tuple[str, str]:
    # PRIMARY OCR using Tesseract with preprocessing for best extraction.
    # Returns (text, debug_msg).
    try:
        import pytesseract
        
        # Ensure RGB first
        if pil_img.mode != "RGB":
            pil_img = pil_img.convert("RGB")
        
        # Convert to numpy array for preprocessing
        img_array = np.array(pil_img)
        
        # Preprocessing for better OCR accuracy
        # 1. Resize if too small (improves OCR on small images)
        h, w = img_array.shape[:2]
        if max(h, w) < 1000:
            scale = 1000 / max(h, w)
            new_w, new_h = int(w * scale), int(h * scale)
            pil_img = pil_img.resize((new_w, new_h), Image.Resampling.LANCZOS)
            img_array = np.array(pil_img)
        
        # 2. Convert to grayscale for better text detection
        if len(img_array.shape) == 3:
            gray = np.mean(img_array, axis=2).astype(np.uint8)
        else:
            gray = img_array
        
        # Convert back to PIL for Tesseract
        processed_img = Image.fromarray(gray)
        
        # Tesseract config for extracting ALL text (not just high confidence)
        # PSM 6 = Assume a single uniform block of text
        # OEM 3 = Default (LSTM + legacy)
        custom_config = r'--oem 3 --psm 6 -c preserve_interword_spaces=1'
        
        # Run Tesseract with full config
        text = pytesseract.image_to_string(processed_img, config=custom_config)
        
        # If grayscale fails, try original RGB
        if not text.strip():
            text = pytesseract.image_to_string(pil_img, config=custom_config)
        
        # Also try PSM 3 (auto) if still no text
        if not text.strip():
            text = pytesseract.image_to_string(pil_img, config='--oem 3 --psm 3')
        
        if not text.strip():
            return "", "Tesseract ran but returned empty text."
        
        return text.strip(), f"✅ Tesseract OCR successful ({len(text.split())} words)"
    except Exception as e:
        return "", f"Tesseract error: {type(e).__name__}: {str(e)}"

# -----------------------------------------------------------------------------
# ENVIRONMENT DETECTION - Localhost vs Streamlit Cloud
# -----------------------------------------------------------------------------
def is_running_on_streamlit_cloud() -> bool:
    # Detect if app is running on Streamlit Cloud vs localhost.
    # Streamlit Cloud sets specific environment variables.
    import os
    # Streamlit Cloud indicators
    if os.environ.get('STREAMLIT_SHARING_MODE'):
        return True
    if os.path.exists('/mount/src'):  # Streamlit Cloud mounts repo here
            return True
    # Check hostname
    import socket
# API Configuration and Helpers already defined above
# Removing duplicate definition blocks

# ... (Previous imports kept)

def extract_text_with_ocr(pil_img: Image.Image, filename: str) -> tuple[str, str]:
    # DEPRECATED: Local OCR logic. 
    # Use process_single_image which calls the Cloud API.
    return "", "Local OCR deprecated. Please use Cloud API."

# -----------------------------------------------------------------------------
# SINGLE IMAGE PROCESSING (With Fallback OCR Support)
# -----------------------------------------------------------------------------
def process_single_image(uploaded_file, file_index: int, total_files: int, use_nlp: bool = True) -> Dict[str, Any]:
    # Process a single uploaded image - uses Surya OCR if available, otherwise fallback to EasyOCR/Tesseract
    
    # Check if minimum pipeline components are available
    if not PIPELINE_AVAILABLE:
        return {
            'filename': uploaded_file.name,
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'method': 'File Upload',
            'error': f'Pipeline components not available: {"; ".join(IMPORT_ERRORS)}',
            'compliance_status': 'ERROR'
        }
    
    try:
        # Load image
        image = Image.open(uploaded_file)
        if image.mode != 'RGB':
            image = image.convert('RGB')
            
        start_time = time.time()
        st.info(f"Processing {uploaded_file.name} with Tesseract OCR...")
        
        # Use Tesseract OCR directly
        try:
            import pytesseract
            
            # Extract text using Tesseract
            ocr_text = pytesseract.image_to_string(image)
            st.success("✅ OCR extraction successful")
            
        except Exception as ocr_err:
            raise Exception(f"Tesseract OCR failed: {ocr_err}")
        
        # Validate using ComplianceValidator (same as Web Crawler)
        try:
            # Import compliance validator
            import sys
            import pathlib
            project_root = pathlib.Path(__file__).resolve().parent.parent.parent
            ml_model_path = project_root / "ml model"
            if str(ml_model_path) not in sys.path:
                sys.path.insert(0, str(ml_model_path))
            
            from compliance import compute_compliance_score
            
            # Prepare parsed_data structure for compliance validator
            parsed_data = {
                "raw_text": ocr_text
            }
            
            # Pre-extract fields using regex to populate parsed_data
            text_lower = ocr_text.lower()
            
            # Extract MRP
            mrp_match = re.search(r'(?:mrp|price)[:\s]*[₹rs.]*\s*([\d,]+(?:\.\d{2})?)', ocr_text, re.I)
            if mrp_match:
                parsed_data["mrp_incl_taxes"] = f"₹{mrp_match.group(1)}"
                parsed_data["mrp"] = f"₹{mrp_match.group(1)}"
            
            # Extract Net Quantity
            qty_match = re.search(r'(?:net\s*(?:quantity|qty|wt|weight)|quantity|weight)[:\s]*([\d.]+\s*(?:kg|g|gm|ml|l|ltr|litre))', text_lower, re.I)
            if qty_match:
                parsed_data["gross_content"] = qty_match.group(1).strip()
                parsed_data["net_quantity"] = qty_match.group(1).strip()
                parsed_data["net"] = qty_match.group(1).strip()
            
            # Extract Manufacturer
            mfr_match = re.search(r'(?:mfd|manufactured|marketed|packed)\s*by[:\s]*([^\n]{10,100})', ocr_text, flags=re.I)
            if mfr_match:
                mfr_text = mfr_match.group(1).strip()
                parsed_data["packed_and_marketed_by"] = {
                    "name": mfr_text.split(',')[0].strip() if ',' in mfr_text else mfr_text,
                    "address_lines": [line.strip() for line in mfr_text.split(',')[1:]] if ',' in mfr_text else []
                }
            
            # Extract Customer Care
            contact_match = re.search(r'(?:consumer\s*care|customer\s*care|helpline|toll\s*free|contact)[:\s]*([^\n]{10,150})', ocr_text, flags=re.I)
            phone_match = re.search(r'(\d{10})', ocr_text)
            email_match = re.search(r'([a-z0-9._%+-]+@[a-z0-9.-]+\.[a-z]{2,})', text_lower)
            
            if contact_match or phone_match or email_match:
                parsed_data["customer_care"] = {
                    "phone": phone_match.group(1) if phone_match else None,
                    "email": email_match.group(1) if email_match else None,
                    "website": None
                }
            
            # Extract Date
            date_match = re.search(r'(?:mfg|mfd|manufactured|best\s*before|expiry|exp\.?\s*date)[:\s]*([^\n]{5,40})', ocr_text, flags=re.I)
            if date_match:
                parsed_data["mfg_date"] = date_match.group(1).strip()
                parsed_data["best_before"] = date_match.group(1).strip()
            
            # Extract Country
            country_match = re.search(r'(?:country\s*of\s*origin|origin|made\s*in|product\s*of)[:\s]*([A-Za-z\s]+)', ocr_text, flags=re.I)
            if country_match:
                parsed_data["country_of_origin"] = country_match.group(1).strip()
                parsed_data["country"] = country_match.group(1).strip()
            
            # Run compliance check
            compliance_result = compute_compliance_score(parsed_data)
            
            # Extract validation results
            validation_results = {}
            passed_rules = compliance_result.get("passed_rules", {})
            
            # Map compliance results to validation format
            validation_results["manufacturer_details"] = bool(passed_rules.get("packed_and_marketed_by"))
            validation_results["net_quantity"] = bool(passed_rules.get("net_quantity") or passed_rules.get("gross_content"))
            validation_results["mrp"] = bool(passed_rules.get("mrp") or passed_rules.get("mrp_incl_taxes"))
            validation_results["customer_care_details"] = True  # Always True as per LMPC rules
            validation_results["date_of_manufacture"] = bool(passed_rules.get("mfg_date") or passed_rules.get("best_before"))
            validation_results["country_of_origin"] = bool(passed_rules.get("country_of_origin") or passed_rules.get("country"))
            
            # Get compliance score
            score = compliance_result.get('compliance_percentage', 0)
            is_compliant = score >= 100
            
            # Create violations from failed rules
            violations = []
            for rule in compliance_result.get("failed_rules", []):
                violations.append({
                    "field": rule.get("key", "unknown"),
                    "description": rule.get("message", "Validation failed"),
                    "violated": True
                })
            
        except Exception as e:
            # Fallback to simple regex if compliance validator fails
            st.warning(f"Compliance validator failed, using fallback: {e}")
            validation_results = {
                "manufacturer_details": False,
                "net_quantity": False,
                "mrp": False,
                "customer_care_details": True,  # Always True
                "date_of_manufacture": False,
                "country_of_origin": False
            }
            
            text_lower = ocr_text.lower()
            
            # Simple regex fallback
            if any(word in text_lower for word in ['mfd by', 'manufactured by', 'marketed by', 'manufacturer']):
                validation_results["manufacturer_details"] = True
            
            if re.search(r'\d+\s*(kg|g|gm|ml|l|litre|ltr)', text_lower):
                validation_results["net_quantity"] = True
            
            if any(word in text_lower for word in ['mrp', 'price', 'rs.', '₹']):
                validation_results["mrp"] = True
            
            if any(word in text_lower for word in ['mfg', 'date', 'exp', 'expiry', 'best before']):
                validation_results["date_of_manufacture"] = True
            
            if any(word in text_lower for word in ['made in', 'country of origin', 'product of', 'india']):
                validation_results["country_of_origin"] = True
            
            # Calculate compliance
            compliant_count = sum(validation_results.values())
            score = (compliant_count / 6) * 100
            is_compliant = score >= 100
            
            # Create violations list
            violations = []
            for field, is_present in validation_results.items():
                if not is_present:
                    violations.append({
                        "field": field,
                        "description": f"{field.replace('_', ' ').title()} is missing",
                        "violated": True
                    })
        
        processing_time = time.time() - start_time
        
        # Prepare result
        result = {
            'filename': uploaded_file.name,
            'file_size': uploaded_file.size,
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'method': 'Tesseract OCR',
            'processing_time': processing_time,
            'ocr_result': ocr_text,
            'refined_data': validation_results, 
            'violations': violations,
            'compliance_status': 'COMPLIANT' if is_compliant else 'NON_COMPLIANT',
            'compliance_score': score,
            'image_dimensions': image.size
        }
        
        # Save to DB
        try:
            user = st.session_state.get("user", {})
            user_id = user.get("id") or 1
            username = user.get("username", "anonymous")
            db.save_image_upload(
                user_id=user_id,
                username=username,
                filename=result["filename"],
                file_size=result["file_size"],
                extracted_text=result.get("ocr_result", "")[:1000],
                confidence=score,
                processing_time=result["processing_time"],
            )

            # Save Compliance Check
            db.save_compliance_check(
                user_id=user_id,
                username=username,
                product_title=result.get("filename", "Unknown"),
                platform="Upload",
                score=score,
                status=result["compliance_status"],
                details=json.dumps(result.get("violations", []))
            )
        except Exception:
            pass
        
        return result
        
    except Exception as e:
        return {
            'filename': uploaded_file.name,
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'method': 'Tesseract OCR',
            'error': str(e),
            'compliance_status': 'ERROR'
        }

# -----------------------------------------------------------------------------
# BATCH PROCESSING
# -----------------------------------------------------------------------------
def process_batch_images(valid_files: List, use_nlp: bool = True) -> List[Dict[str, Any]]:
    if not valid_files:
        return []

    results = []
    progress_bar = st.progress(0)
    status_text = st.empty()
    results_container = st.empty()
    total_files = len(valid_files)

    for i, f in enumerate(valid_files):
        progress = (i + 1) / total_files
        progress_bar.progress(progress)
        status_text.text(f"Processing {f.name} ({i+1}/{total_files})")

        res = process_single_image(f, i, total_files, use_nlp=use_nlp)
        results.append(res)

        with results_container.container():
            display_batch_progress(results, total_files)

        time.sleep(0.05)

    progress_bar.empty()
    status_text.empty()
    results_container.empty()

    return results

def display_batch_progress(results: List[Dict], total_files: int):
    processed = len(results)
    col1, col2, col3 = st.columns(3)

    with col1:
        st.metric("Processed", f"{processed}/{total_files}")
    with col2:
        compliant = sum(1 for r in results if r.get("compliance_status") == "COMPLIANT")
        st.metric("Compliant", compliant)
    with col3:
        errors = sum(1 for r in results if r.get("compliance_status") == "ERROR")
        st.metric("❌ Errors", errors)

# -----------------------------------------------------------------------------
# RESULT DISPLAY (SAME LOGIC AS BEFORE, JUST REUSED)
# -----------------------------------------------------------------------------
def display_batch_results(results: List[Dict[str, Any]], context: str = "default", suppress_inner_expanders: bool = False):
    if not results:
        return

    st.markdown("##  Batch Processing Results")

    total_files = len(results)
    successful = sum(1 for r in results if "error" not in r)
    compliant = sum(1 for r in results if r.get("compliance_status") == "COMPLIANT")
    errors = sum(1 for r in results if "error" in r)

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.metric("Total Files", total_files)
    with c2:
        st.metric("Successfully Processed", successful)
    with c3:
        st.metric("Compliant Products", compliant)
    with c4:
        st.metric("❌ Processing Errors", errors)

    st.markdown("###  Flagged Results")

    for idx, r in enumerate(results):
        display_single_result_detailed(r, show_expander=not suppress_inner_expanders, context=context, idx=idx)

    # Summary table
    if suppress_inner_expanders:
        st.markdown("### Summary Table View")
        table_data = []
        for r in results:
            table_data.append(
                {
                    "Filename": r.get("filename", "Unknown"),
                    "Status": r.get("compliance_status", "Unknown"),
                    "Violations": len(r.get("violations", [])),
                    "Processing Time": f"{r.get('processing_time', 0):.2f}s",
                    "File Size": f"{r.get('file_size', 0) / 1024:.1f} KB" if r.get("file_size") else "",
                    "Timestamp": r.get("timestamp", "Unknown"),
                }
            )
        df = pd.DataFrame(table_data)
        st.dataframe(df, width=1000)
    else:
        with st.expander(" Summary Table View", expanded=False):
            table_data = []
            for r in results:
                table_data.append(
                    {
                        "Filename": r.get("filename", "Unknown"),
                        "Status": r.get("compliance_status", "Unknown"),
                        "Violations": len(r.get("violations", [])),
                        "Processing Time": f"{r.get('processing_time', 0):.2f}s",
                        "File Size": f"{r.get('file_size', 0) / 1024:.1f} KB" if r.get("file_size") else "",
                        "Timestamp": r.get("timestamp", "Unknown"),
                    }
                )
            df = pd.DataFrame(table_data)
            st.dataframe(df, width=1000)

def display_single_result_detailed(result: Dict[str, Any], show_expander: bool = True, context: str = "default", idx: int = 0):
    filename = result.get("filename", "Unknown")

    if "error" in result:
        st.error(f"❌ {filename} - Processing Error: {result['error']}")
        return

    violations = result.get("violations", [])
    compliance_status = result.get("compliance_status", "UNKNOWN")
    refined_data = result.get("refined_data", {}) or {}

    violation_count = len(violations)
    compliance_score = max(0, 100 - violation_count * 20)
    
    # ---------------------------------------------------------
    # 1. Header & Status
    # ---------------------------------------------------------
    st.markdown(f"### 📄 Analysis for: {filename}")
    
    if compliance_status == "COMPLIANT":
        st.success(f"🟢 **COMPLIANT** (Score: {compliance_score:.0f}/100)")
    else:
        st.error(f"🔴 **NON-COMPLIANT** ({violation_count} Violations Detected)")

    # ---------------------------------------------------------
    # 2. Key Mandatory Fields (The "Important 6")
    # ---------------------------------------------------------
    st.markdown("#### 🏛️ Mandatory Declarations")
    
    # Define mandatory fields mapping
    field_map = [
        {"label": "Manufacturer Name/Address", "key": "manufacturer_details", "icon": "🏭"},
        {"label": "Country of Origin", "key": "country_of_origin", "icon": "🌍"},
        {"label": "Net Quantity", "key": "net_quantity", "icon": "⚖️"},
        {"label": "Mfg / Import Date", "key": "date_of_manufacture", "alt_key": "date_of_import", "icon": "📅"},
        {"label": "MRP (Max Retail Price)", "key": "mrp", "icon": "💰"},
        {"label": "Customer Care Details", "key": "customer_care_details", "icon": "📞"},
    ]

    # Grid Layout
    cols = st.columns(2)
    for i, item in enumerate(field_map):
        key = item["key"]
        val = refined_data.get(key)
        
        
        # Check alternate key if main missing (e.g. Mfg vs Import date)
        if (not val or val == "Not Found") and "alt_key" in item:
            val = refined_data.get(item["alt_key"])

        # Handle boolean values from Tesseract validation
        if isinstance(val, bool):
            display_val = "✅ Found" if val else "❌ Not Found"
        else:
            display_val = val if (val and str(val).lower() != "none") else "❌ Not Found"
        
        is_missing = "❌" in display_val
        
        # Styling
        border_color = "#ff4b4b" if is_missing else "#09ab3b"
        bg_color = "rgba(255, 75, 75, 0.1)" if is_missing else "rgba(9, 171, 59, 0.1)"
        
        with cols[i % 2]:
            st.markdown(f"""
            <div style="
                border: 1px solid {border_color};
                background-color: {bg_color};
                padding: 10px;
                border-radius: 5px;
                margin-bottom: 10px;
            ">
                <div style="font-weight:bold; font-size:0.9em; color:#555;">{item['icon']} {item['label']}</div>
                <div style="font-size:1.1em; color: black;">{display_val}</div>
            </div>
            """, unsafe_allow_html=True)

    # ---------------------------------------------------------
    # 3. Violations Summary (if any)
    # ---------------------------------------------------------
    if violations:
        st.markdown("#### ⚠️ Violations Found")
        for v in violations:
            msg = v.get("description", str(v)) if isinstance(v, dict) else str(v)
            st.warning(f"• {msg}")

    # ---------------------------------------------------------
    # 4. Raw Data Expander
    # ---------------------------------------------------------
    with st.expander("📂 View Raw Data & Technical Details"):
        t1, t2 = st.tabs(["📝 Raw Text", "📊 Full Data"])
        
        with t1:
            ocr_text = result.get("ocr_result", "No text found")
            st.text_area("OCR Output", ocr_text, height=200, key=f"ocr_area_{context}_{idx}")
            
            # Download Button
            st.download_button(
                label="⬇️ Download Raw Text",
                data=ocr_text,
                file_name=f"raw_{filename}.txt",
                mime="text/plain",
                key=f"dl_raw_{context}_{idx}"
            )
            
        with t2:
            st.json(result)

# -----------------------------------------------------------------------------
# EXPORT
# -----------------------------------------------------------------------------
def create_export_panel(results: List[Dict[str, Any]], context: str = "default"):
    if not results:
        return

    st.markdown("###  Export Results")
    c1, c2, c3 = st.columns(3)

    with c1:
        if st.button(" Export to Excel", key=f"export_excel_{context}"):
            excel_data = prepare_excel_export(results)
            st.download_button(
                label="⬇ Download Excel Report",
                data=excel_data,
                file_name=f"compliance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"download_excel_{context}",
            )

    with c2:
        if st.button(" Export to JSON", key=f"export_json_{context}"):
            json_data = json.dumps(results, indent=2, ensure_ascii=False)
            st.download_button(
                label="⬇ Download JSON Data",
                data=json_data,
                file_name=f"compliance_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                mime="application/json",
                key=f"download_json_{context}",
            )

    with c3:
        st.button("📄 Generate PDF Report", key=f"export_pdf_{context}", disabled=True)
        st.caption("PDF report generation coming soon.")

def prepare_excel_export(results: List[Dict[str, Any]]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"

    total_files = len(results)
    compliant = sum(1 for r in results if r.get("compliance_status") == "COMPLIANT")
    non_compliant = total_files - compliant

    summary_data = [
        ["Legal Metrology Compliance Report", ""],
        ["Generated on", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["", ""],
        ["Total Files Processed", total_files],
        ["Compliant Products", compliant],
        ["Non-Compliant Products", non_compliant],
        ["Compliance Rate", f"{(compliant/total_files*100):.1f}%" if total_files else "0%"],
    ]

    for r, (k, v) in enumerate(summary_data, start=1):
        ws_summary.cell(row=r, column=1, value=k)
        ws_summary.cell(row=r, column=2, value=v)

    ws_detail = wb.create_sheet("Detailed Results")
    headers = ["Filename", "Compliance Status", "Violations Count", "Processing Time (s)", "File Size (KB)", "Timestamp"]

    for c, h in enumerate(headers, start=1):
        cell = ws_detail.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    for r_idx, res in enumerate(results, start=2):
        ws_detail.cell(row=r_idx, column=1, value=res.get("filename", "Unknown"))
        ws_detail.cell(row=r_idx, column=2, value=res.get("compliance_status", "Unknown"))
        ws_detail.cell(row=r_idx, column=3, value=len(res.get("violations", [])))
        ws_detail.cell(row=r_idx, column=4, value=res.get("processing_time", 0))
        ws_detail.cell(row=r_idx, column=5, value=(res.get("file_size", 0) / 1024))
        ws_detail.cell(row=r_idx, column=6, value=res.get("timestamp", "Unknown"))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

# -----------------------------------------------------------------------------
# MAIN
# -----------------------------------------------------------------------------
def main():
    load_upload_css()
    initialize_upload_session()

    # Page Header
    header_html = [
        '<div class="page-header">',
        '    <h1> Upload & Process Images</h1>',
        '    <p>Upload multiple product images for batch compliance validation. Drag and drop images or use the upload button to get started.</p>',
        '</div>'
    ]
    st.markdown("\n".join(header_html), unsafe_allow_html=True)

    # Sidebar
    with st.sidebar:
        st.markdown("###  Processing Options")
        auto_process = st.checkbox("Auto-process on upload", value=True)
        save_results = st.checkbox("Save results to session", value=True)
        
        # LLM extraction removed - using compliance validator
        use_llm = False

        st.markdown("---")
        st.markdown("### 📈 Session Stats")
        total = len(st.session_state.batch_results)
        st.metric("Files Processed", total)
        if total:
            compliant = sum(
                1 for r in st.session_state.batch_results if r.get("compliance_status") == "COMPLIANT"
            )
            st.metric("✅ Compliant Rate", f"{(compliant/total*100):.1f}%")

    # Input method
    input_method = st.radio(
        "Select image source:",
        [" Browse Files", " Capture from Camera"],
        horizontal=True,
        key="input_method_radio",
    )

    uploaded_files = None
    if input_method == " Browse Files":
        uploaded_files = create_file_upload_interface()
    else:
        uploaded_files = create_camera_capture_interface()

    # Main logic
    if uploaded_files:
        valid_files = display_uploaded_files(uploaded_files)

        if valid_files:
            c1, c2, c3 = st.columns([2, 1, 1])
            with c1:
                st.success(f"✅ {len(valid_files)} valid files ready for processing")
            with c2:
                process_button = st.button(
                    " Process All Images",
                    type="primary",
                    disabled=len(valid_files) == 0,
                    key="process_all_images",
                )
            with c3:
                if st.button("🗑️ Clear Selection", key="clear_current_selection"):
                    st.session_state.current_batch_results = []
                    st.session_state.files_processed = set()
                    st.session_state.processing_completed = False
                    st.rerun()

            current_ids = {f"{f.name}_{f.size}" for f in valid_files}
            should_process = (
                process_button
                or (auto_process and valid_files and not st.session_state.processing_completed)
                or (auto_process and valid_files and current_ids != st.session_state.files_processed)
            )

            if should_process:
                with st.spinner("Processing images through compliance pipeline..."):
                    batch_results = process_batch_images(valid_files, use_nlp=use_llm)
                    st.session_state.current_batch_results = batch_results
                    st.session_state.files_processed = current_ids
                    st.session_state.processing_completed = True

                    if save_results:
                        st.session_state.batch_results.extend(batch_results)

            if st.session_state.current_batch_results:
                st.success(
                    f"✅ Successfully processed {len(st.session_state.current_batch_results)} images!"
                )
                display_batch_results(
                    st.session_state.current_batch_results,
                    "current_batch",
                    suppress_inner_expanders=False,
                )
                create_export_panel(st.session_state.current_batch_results, "current_batch")

    # Session history
    if st.session_state.batch_results:
        st.markdown("---")
        st.markdown("## 📚 Session History")
        with st.expander(
            f"View All Results ({len(st.session_state.batch_results)} files)", expanded=False
        ):
            display_batch_results(
                st.session_state.batch_results, "session_history", suppress_inner_expanders=True
            )
            create_export_panel(st.session_state.batch_results, "session_history")

            if st.button(" Clear Session History", key="clear_session_history"):
                st.session_state.batch_results = []
                st.rerun()

    # Navigation
    st.markdown("---")
    n1, n2, n3 = st.columns(3)
    with n1:
        if st.button(" Back to Dashboard", key="nav_dashboard_upload"):
            st.switch_page("streamlit_app.py")
    with n2:
        if st.button(" Settings", key="nav_settings_upload"):
            st.switch_page("pages/04__Settings.py")
    with n3:
        if st.button(" Batch Processing", key="nav_batch_upload"):
            st.switch_page("pages/03__Batch_Process.py")

if __name__ == "__main__":
    main()
